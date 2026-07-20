"""Excel preview endpoint — reads .xlsx/.xls/.ods via local fs or SFTP,
parses with openpyxl, and returns structured sheet data for the frontend table.

Only the first MAX_ROWS rows of each sheet are returned to avoid sending
enormous payloads. Cell values are coerced to strings; None cells become "".
"""
from __future__ import annotations

import asyncio
import io
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel

from ..config import AppConfig
from ..ssh_pool import get_pool
from ..validation import PathValidator, ValidationError
from .dependencies import APIDependencies
from .helpers import _local_read_bytes, _read_file_bytes

MAX_ROWS = 2000
MAX_COLS = 200
# Hard cap on raw file bytes we'll parse — 50 MB is generous for a spreadsheet
MAX_FILE_BYTES = 50_000_000


class ExcelSheet(BaseModel):
    name: str
    rows: list[list[str]]
    truncated_rows: bool = False
    truncated_cols: bool = False


class ExcelPreviewResponse(BaseModel):
    sheets: list[ExcelSheet]
    active_sheet: str
    file_too_large: bool = False


def _parse_workbook(data: bytes) -> ExcelPreviewResponse:
    """Parse raw workbook bytes with openpyxl. Runs in a thread."""
    import openpyxl  # local import — only load when actually needed

    wb = openpyxl.load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=True,   # return cached values, not formulas
    )

    sheets: list[ExcelSheet] = []
    active_name = wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else "")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        truncated_rows = False
        truncated_cols = False

        for row in ws.iter_rows():
            if len(rows) >= MAX_ROWS:
                truncated_rows = True
                break
            cells: list[str] = []
            for cell in row:
                if len(cells) >= MAX_COLS:
                    truncated_cols = True
                    break
                val = cell.value
                if val is None:
                    cells.append("")
                else:
                    cells.append(str(val))
            rows.append(cells)

        # Strip trailing empty rows
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()

        sheets.append(ExcelSheet(
            name=sheet_name,
            rows=rows,
            truncated_rows=truncated_rows,
            truncated_cols=truncated_cols,
        ))

    wb.close()
    return ExcelPreviewResponse(sheets=sheets, active_sheet=active_name)


def get_router(deps: APIDependencies) -> APIRouter:
    router = APIRouter()

    @router.get("/boxes/{name}/excel", response_model=ExcelPreviewResponse)
    async def api_excel_preview(
        name: str,
        path: str = Query(...),
        application_config: AppConfig = Depends(deps.get_application_config),
    ) -> Any:
        box = deps.get_box_or_404(application_config, name)

        if box.transport == "local":
            try:
                validated = PathValidator.validate_local_path(path)
            except ValidationError as exc:
                raise HTTPException(status_code=400, detail=str(exc))

            data, too_large = await _local_read_bytes(validated, MAX_FILE_BYTES)
            if too_large:
                return ExcelPreviewResponse(sheets=[], active_sheet="", file_too_large=True)
        else:
            try:
                validated = PathValidator.validate_remote_path(path)
            except ValidationError as exc:
                raise HTTPException(status_code=400, detail=str(exc))

            ssh_pool = get_pool()
            try:
                async with ssh_pool.connection(
                    box, lambda: deps.connect_for_box(box, application_config)
                ) as conn:
                    data, too_large = await _read_file_bytes(conn, validated, MAX_FILE_BYTES)
            except Exception as exc:
                raise HTTPException(status_code=502, detail=f"SSH error: {exc}")

            if too_large:
                return ExcelPreviewResponse(sheets=[], active_sheet="", file_too_large=True)

        try:
            result = await asyncio.to_thread(_parse_workbook, data)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not parse workbook: {exc}")

        return result

    return router
